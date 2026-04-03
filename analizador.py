# =============================================================
# ANALIZADOR DE CORPUS DE AMPAROS
# Versión 1.0 — Investigación Doctoral en Derecho Empírico
# =============================================================
#
# Mejoras sobre la arquitectura original:
#  1. Modelo fijado como constante fácil de cambiar
#  2. Pausa propagada al interior de analizar_sentencia (respeta pausa entre reintentos)
#  3. UI thread-safe: el hilo de fondo NUNCA toca session_state;
#     la interfaz monitorea con polling a través de threading.Event
#  4. Archivos .bat + INSTRUCCIONES.txt incluidos para usuario no programador
# =============================================================

import streamlit as st
import anthropic
from anthropic import RateLimitError, APIConnectionError, BadRequestError
import json
import re
import hashlib
import threading
import time
import io
import os
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# =============================================================
# CONSTANTES GLOBALES
# =============================================================

# ⚠️  Si necesita cambiar el modelo, modifique SOLO esta línea.
MODELO = "claude-sonnet-4-5"

MAX_WORKERS      = 5    # llamadas simultáneas
MAX_INTENTOS     = 6    # reintentos por ERRORES REALES (JSON inválido, conexión rota)
                        # Los errores 429 (límite de API) se reintentan indefinidamente
SPACING_API      = 9.0  # segundos mínimos entre dos llamadas API consecutivas
                        # 40k TPM / ~5500 tokens/doc ≈ 7 docs/min → 1 cada 8.5s
MAX_CHARS_TEXTO  = 80_000  # ~20 000 tokens; documentos más largos se truncan
ARCHIVO_PROGRESO = "progreso.json"
ARCHIVO_BITACORA = "bitacora.json"
ARCHIVO_EXCEL    = "resultados_amparos.xlsx"
ARCHIVO_ESTADO   = "estado.json"  # comunica estado del hilo a la UI

PATRONES_LISTA = [
    "p1_omision",
    "p2_precariedad",
    "p3_documento_digital",
    "p4_sobreseimiento",
    "p5_no_transcripcion",
]

PATRONES_DESCRIPCION = {
    "p1_omision": (
        "P1 — Omisión administrativa: La sentencia tiene origen en que la autoridad no respondió "
        "un escrito o no cumplió una sentencia previa. El amparo se interpone por omisión de "
        "respuesta o negativa ficta."
    ),
    "p2_precariedad": (
        "P2 — Cláusula de precariedad: El juzgador inserta en la resolución una justificación de "
        "la falta de integración física del expediente por insuficiencia de insumos materiales "
        "básicos del juzgado (tóner, hojas, papel)."
    ),
    "p3_documento_digital": (
        "P3 — Rechazo de documento digital: El juzgador descalifica o rechaza documentos "
        "electrónicos oficiales por no cumplir con requisitos propios del documento en papel: "
        "falta de firma manuscrita, falta de sello físico, o los trata como copias simples sin "
        "valor probatorio."
    ),
    "p4_sobreseimiento": (
        "P4 — Sesgo de anclaje en sobreseimiento: En una resolución de sobreseimiento, el texto "
        "no muestra ponderación activa de los argumentos del quejoso frente a la negativa "
        "categórica de la autoridad responsable. El sobreseimiento parece derivar automáticamente "
        "del informe previo de la autoridad."
    ),
    "p5_no_transcripcion": (
        "P5 — No transcripción de conceptos de violación: La sentencia aplica explícitamente la "
        "práctica de no transcribir los conceptos de violación o agravios del quejoso, invocando "
        "que es innecesaria su reproducción para cumplir con congruencia y exhaustividad."
    ),
}

# El prompt NO contiene llaves de Python salvo {texto}, que se reemplaza con .replace()
# para evitar colisiones con el contenido de las sentencias.
PROMPT_BASE = """Eres un asistente jurídico especializado en análisis de sentencias de amparo mexicanas.

Analiza el siguiente texto de una sentencia y determina si contiene cada uno de los cinco patrones jurídicos descritos a continuación.

Para cada patrón indica:
- "presente": true si el patrón está claramente presente en el texto, false si no lo está
- "fragmento": el fragmento textual LITERAL de la sentencia donde aparece el patrón (copiado exactamente del texto, sin modificar ni parafrasear), o "" si el patrón no está presente

PATRONES A DETECTAR:

P1 — OMISIÓN ADMINISTRATIVA
La sentencia tiene origen en que la autoridad no respondió un escrito o no cumplió una sentencia previa. El amparo se interpone por omisión de respuesta o negativa ficta.

P2 — CLÁUSULA DE PRECARIEDAD
El juzgador inserta en la resolución una justificación de la falta de integración física del expediente por insuficiencia de insumos materiales básicos del juzgado (tóner, hojas, papel).

P3 — RECHAZO DE DOCUMENTO DIGITAL
El juzgador descalifica o rechaza documentos electrónicos oficiales por no cumplir con requisitos propios del documento en papel: falta de firma manuscrita, falta de sello físico, o los trata como copias simples sin valor probatorio.

P4 — SESGO DE ANCLAJE EN SOBRESEIMIENTO
En una resolución de sobreseimiento, el texto no muestra ponderación activa de los argumentos del quejoso frente a la negativa categórica de la autoridad responsable. El sobreseimiento parece derivar automáticamente del informe previo de la autoridad.

P5 — NO TRANSCRIPCIÓN DE CONCEPTOS DE VIOLACIÓN
La sentencia aplica explícitamente la práctica de no transcribir los conceptos de violación o agravios del quejoso, invocando que es innecesaria su reproducción para cumplir con congruencia y exhaustividad.

INSTRUCCIONES IMPORTANTES:
- El fragmento debe ser texto LITERAL de la sentencia, nunca una paráfrasis generada por ti
- Si el patrón no está presente, el fragmento debe ser exactamente ""
- Responde ÚNICAMENTE con el objeto JSON, sin texto adicional antes ni después, sin bloques de código markdown
- Si el texto no es una sentencia de amparo analizable, marca todos los patrones como false

Formato de respuesta requerido (sin variaciones):
{
  "p1_omision":          {"presente": true/false, "fragmento": "..."},
  "p2_precariedad":      {"presente": true/false, "fragmento": "..."},
  "p3_documento_digital":{"presente": true/false, "fragmento": "..."},
  "p4_sobreseimiento":   {"presente": true/false, "fragmento": "..."},
  "p5_no_transcripcion": {"presente": true/false, "fragmento": "..."}
}

SENTENCIA A ANALIZAR:
---
{texto}
---"""


# =============================================================
# MÓDULO 2 — PREPROCESAMIENTO
# =============================================================

def limpiar_texto(texto: str) -> str:
    """
    Limpia el texto de una sentencia eliminando:
    1. El bloque de firma criptográfica al final (FIRMANTE / FIRMA / OCSP / TSP).
    2. Las marcas de versión pública intercaladas (PJF - Versión Pública + bloque).
    3. Líneas de timestamp del firmante (dd/mm/aa hh:mm:ss).
    4. Líneas que son cadenas hexadecimales separadas por espacios.
    No modifica el archivo original.
    """
    lineas   = texto.split("\n")
    resultado = []
    i = 0

    while i < len(lineas):
        linea       = lineas[i]
        linea_strip = linea.strip()

        # ── 1. Cortar en el bloque FIRMANTE (firma criptográfica) ──────────────
        if linea_strip == "FIRMANTE":
            # Buscar "Nombre:" en las próximas 4 líneas para confirmar que es el bloque crypto
            for j in range(i + 1, min(i + 5, len(lineas))):
                if lineas[j].strip().startswith("Nombre:") and "Validez:" in lineas[j]:
                    return "\n".join(resultado).strip()
            # No era el bloque crypto: conservar
            resultado.append(linea)
            i += 1
            continue

        # ── 2. Eliminar marca "PJF - Versión Pública" y su bloque ─────────────
        if linea_strip == "PJF - Versión Pública":
            saltar_hasta = i + 1
            j = i + 1
            pasos = 0
            while j < len(lineas) and pasos < 6:
                sig = lineas[j].strip()
                if sig == "":
                    saltar_hasta = j + 1
                    j += 1; pasos += 1; continue
                # Nombre en mayúsculas (letras, espacios, acentos)
                if re.match(r"^[A-ZÁÉÍÓÚÜÑ][A-ZÁÉÍÓÚÜÑ\s]{3,}$", sig):
                    saltar_hasta = j + 1
                    j += 1; pasos += 1; continue
                # Cadena hex compacta (sin espacios)
                if re.match(r"^[0-9a-f]{20,}$", sig.lower()):
                    saltar_hasta = j + 1
                    j += 1; pasos += 1; continue
                # Timestamp dd/mm/aa hh:mm:ss
                if re.match(r"^\d{2}/\d{2}/\d{2}\s+\d{2}:\d{2}:\d{2}", sig):
                    saltar_hasta = j + 1
                    j += 1; pasos += 1; continue
                break
            i = saltar_hasta
            continue

        # ── 3. Eliminar timestamps sueltos ────────────────────────────────────
        if re.match(r"^\s*\d{2}/\d{2}/\d{2}\s+\d{2}:\d{2}:\d{2}", linea):
            i += 1
            continue

        # ── 4. Eliminar líneas hex separadas por espacios (fragmentos de firma) ─
        if re.match(r"^\s*([0-9a-f]{2}\s+){5,}[0-9a-f]{2}\s*$", linea.lower()):
            i += 1
            continue

        resultado.append(linea)
        i += 1

    return "\n".join(resultado).strip()


# =============================================================
# MÓDULO 3 — ANÁLISIS CON IA
# =============================================================

def construir_prompt(texto_limpio: str) -> str:
    """Inserta el texto en el prompt fijo usando .replace() para evitar conflictos de formato."""
    return PROMPT_BASE.replace("{texto}", texto_limpio)


def _validar_estructura(datos: dict) -> bool:
    """Verifica que el JSON tenga los cinco patrones con las claves requeridas."""
    for patron in PATRONES_LISTA:
        if patron not in datos:
            return False
        if not isinstance(datos[patron], dict):
            return False
        if "presente" not in datos[patron] or "fragmento" not in datos[patron]:
            return False
    return True


def parsear_respuesta(respuesta: str) -> Optional[dict]:
    """
    Extrae y valida el JSON de la respuesta.
    Maneja: JSON puro, JSON dentro de bloques markdown, JSON embebido en texto.
    """
    texto = respuesta.strip()

    # Intento 1: JSON puro
    try:
        datos = json.loads(texto)
        if _validar_estructura(datos):
            return datos
    except json.JSONDecodeError:
        pass

    # Intento 2: JSON dentro de bloque ```json ... ```
    m = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", texto)
    if m:
        try:
            datos = json.loads(m.group(1))
            if _validar_estructura(datos):
                return datos
        except json.JSONDecodeError:
            pass

    # Intento 3: primer objeto JSON que contenga p1_omision
    m = re.search(r"\{[\s\S]*?\"p1_omision\"[\s\S]*?\}", texto)
    if m:
        try:
            datos = json.loads(m.group(0))
            if _validar_estructura(datos):
                return datos
        except json.JSONDecodeError:
            pass

    return None


def _resultado_error(nombre: str) -> dict:
    """Devuelve un resultado vacío marcado como error."""
    return {
        "_archivo":            nombre,
        "_procesado":          True,
        "_error":              True,
        "p1_omision":          {"presente": False, "fragmento": ""},
        "p2_precariedad":      {"presente": False, "fragmento": ""},
        "p3_documento_digital":{"presente": False, "fragmento": ""},
        "p4_sobreseimiento":   {"presente": False, "fragmento": ""},
        "p5_no_transcripcion": {"presente": False, "fragmento": ""},
    }


def _reservar_slot_api() -> None:
    """
    Espaciador global: garantiza al menos SPACING_API segundos entre llamadas.

    FIX: El lock se libera ANTES de dormir, así cada worker calcula su propio
    tiempo de espera y lo cumple de forma independiente. Sin esto, los workers
    se bloquean entre sí porque el lock permanece tomado durante el sleep.
    """
    with _lock_api_slot:
        ahora  = time.time()
        espera = max(0.0, SPACING_API - (ahora - _ultimo_ts_api[0]))
        # Reservar el "slot" sumando la espera al timestamp esperado
        _ultimo_ts_api[0] = ahora + espera
    # El lock ya fue liberado — cada worker duerme en paralelo su propio tiempo
    if espera > 0:
        time.sleep(espera)


def analizar_sentencia(
    nombre: str,
    texto: str,
    api_key: str,
    pausa_event: Optional[threading.Event] = None,
) -> dict:
    """
    Analiza una sentencia con la API de Claude.
    - Trunca textos muy largos antes de enviar (evita error 400 de token limit).
    - Espaciador global entre llamadas para no saturar el bucket de tokens.
    - Los errores 429 (límite de API) se reintentan indefinidamente con jitter.
      NO se cuentan contra MAX_INTENTOS.
    - Los errores 400 de texto demasiado largo truncan y reintentan sin contar.
    - Solo cuentan como intentos reales: JSON inválido, conexión rota, otros.
    """
    import random

    if pausa_event and pausa_event.is_set():
        return _resultado_error(nombre)

    texto_limpio = limpiar_texto(texto)

    # Truncar si el documento es demasiado largo (evita error 400 "prompt too long")
    truncado = False
    if len(texto_limpio) > MAX_CHARS_TEXTO:
        texto_limpio = texto_limpio[:MAX_CHARS_TEXTO]
        truncado = True

    prompt = construir_prompt(texto_limpio)
    client = anthropic.Anthropic(api_key=api_key)

    intentos_reales = 0

    while intentos_reales < MAX_INTENTOS:
        if pausa_event and pausa_event.is_set():
            return _resultado_error(nombre)

        try:
            _reservar_slot_api()
            if truncado:
                escribir_estado(f"Analizando (texto truncado): {nombre}")
            else:
                escribir_estado(f"Analizando: {nombre}")

            mensaje = client.messages.create(
                model=MODELO,
                max_tokens=2048,
                messages=[{"role": "user", "content": prompt}],
            )
            texto_respuesta = mensaje.content[0].text
            datos = parsear_respuesta(texto_respuesta)

            if datos is None:
                intentos_reales += 1
                escribir_estado(
                    f"Respuesta no válida ({nombre}), reintentando "
                    f"[{intentos_reales}/{MAX_INTENTOS}]..."
                )
                time.sleep(5 * intentos_reales)
                continue

            escribir_estado(f"✓ Completada: {nombre}")
            return {"_archivo": nombre, "_procesado": True, "_error": False,
                    "_truncado": truncado, **datos}

        except RateLimitError as e:
            # NO cuenta contra intentos_reales — condición temporal de la API
            espera = 60
            try:
                resp = getattr(e, "response", None)
                if resp is not None:
                    ra = (resp.headers.get("retry-after")
                          or resp.headers.get("x-ratelimit-reset-requests"))
                    if ra:
                        espera = max(int(float(ra)), 15)
            except Exception:
                pass
            jitter = random.uniform(2, 20)
            espera_total = espera + jitter
            escribir_estado(
                f"⏳ Límite de API — esperando {espera_total:.0f}s  (normal, no es error)"
            )
            time.sleep(espera_total)

        except BadRequestError as e:
            # Error 400 — puede ser texto demasiado largo incluso después de truncar,
            # o algún problema con el contenido. Intentamos reducir más el texto.
            msg = str(e)
            if "too long" in msg.lower() or "token" in msg.lower():
                # Reducir texto a la mitad y reintentar sin contar como error real
                mitad = len(texto_limpio) // 2
                if mitad > 5000:
                    texto_limpio = texto_limpio[:mitad]
                    prompt = construir_prompt(texto_limpio)
                    truncado = True
                    escribir_estado(
                        f"⚠️ Texto muy largo, reduciendo a {mitad} chars y reintentando: {nombre}"
                    )
                    continue  # reintento sin consumir intentos_reales
                else:
                    # Ya no se puede reducir más
                    intentos_reales += 1
                    escribir_estado(
                        f"Error 400 irreducible en {nombre}: {msg[:120]}"
                    )
            else:
                # Otro tipo de error 400 (API key, formato, etc.)
                intentos_reales += 1
                escribir_estado(
                    f"Error de solicitud en {nombre}: {msg[:150]}  "
                    f"[{intentos_reales}/{MAX_INTENTOS}]"
                )
            time.sleep(5)

        except APIConnectionError:
            intentos_reales += 1
            espera = 10 * intentos_reales
            escribir_estado(
                f"Sin conexión ({nombre}). Reintentando en {espera}s "
                f"[{intentos_reales}/{MAX_INTENTOS}]..."
            )
            time.sleep(espera)

        except Exception as e:
            intentos_reales += 1
            escribir_estado(
                f"Error inesperado ({nombre}): {str(e)[:150]}  "
                f"[{intentos_reales}/{MAX_INTENTOS}]"
            )
            time.sleep(10)

    return _resultado_error(nombre)


# =============================================================
# MÓDULO 4 — ORQUESTADOR
# =============================================================

_lock_progreso  = threading.Lock()
_lock_estado    = threading.Lock()
_lock_api_slot  = threading.Lock()   # espaciador global entre llamadas API
_ultimo_ts_api  = [0.0]              # timestamp de la última llamada iniciada


def escribir_estado(msg: str, inicio: float = None, n_previo: int = None) -> None:
    """Escribe un mensaje de estado que la UI puede leer y mostrar."""
    with _lock_estado:
        try:
            datos: dict = {"mensaje": msg, "ts": time.time()}
            if inicio is not None:
                datos["inicio"] = inicio
            if n_previo is not None:
                datos["n_previo"] = n_previo
            # Preservar 'inicio' y 'n_previo' si ya existen en estado.json
            if os.path.exists(ARCHIVO_ESTADO):
                try:
                    with open(ARCHIVO_ESTADO, "r", encoding="utf-8") as f:
                        prev = json.load(f)
                        if "inicio" not in datos and "inicio" in prev:
                            datos["inicio"] = prev["inicio"]
                        if "n_previo" not in datos and "n_previo" in prev:
                            datos["n_previo"] = prev["n_previo"]
                except Exception:
                    pass
            with open(ARCHIVO_ESTADO, "w", encoding="utf-8") as f:
                json.dump(datos, f)
        except Exception:
            pass


def leer_estado() -> dict:
    """Lee el estado completo del hilo de fondo."""
    try:
        if os.path.exists(ARCHIVO_ESTADO):
            with open(ARCHIVO_ESTADO, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def guardar_progreso(progreso: dict) -> None:
    """Escribe progreso.json de forma thread-safe."""
    with _lock_progreso:
        with open(ARCHIVO_PROGRESO, "w", encoding="utf-8") as f:
            json.dump(progreso, f, ensure_ascii=False, indent=2)


def cargar_progreso() -> dict:
    """Lee progreso.json o devuelve dict vacío si no existe."""
    if os.path.exists(ARCHIVO_PROGRESO):
        try:
            with open(ARCHIVO_PROGRESO, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def extraer_contexto(texto: str, fragmento: str, ventana: int = 20) -> str:
    """
    Extrae el fragmento con líneas de contexto.
    Si fragmento está en texto, retorna líneas circundantes.
    Si no está, retorna el fragmento tal cual.
    """
    if not fragmento or not texto:
        return fragmento

    lineas = texto.split("\n")

    # Buscar línea que contenga el inicio del fragmento
    idx_inicio = -1
    for i, linea in enumerate(lineas):
        if fragmento[:50] in linea:
            idx_inicio = i
            break

    if idx_inicio == -1:
        return fragmento

    # Extraer ventana (anterior + línea + posterior)
    desde = max(0, idx_inicio - ventana)
    hasta = min(len(lineas), idx_inicio + ventana + 1)

    contexto_lineas = lineas[desde:hasta]
    return "\n".join(contexto_lineas)


def guardar_validacion(nombre_archivo: str, patron: str, estado: str) -> None:
    """
    Guarda el estado de validación de un patrón en progreso.json.
    estado puede ser: "Pendiente", "Confirmado", "Rechazado"
    """
    progreso = cargar_progreso()

    if nombre_archivo not in progreso:
        return

    clave_validacion = f"_validacion_{patron}"
    progreso[nombre_archivo][clave_validacion] = estado

    guardar_progreso(progreso)


def correr_analisis(
    archivos_dict: dict,
    api_key: str,
    pausa_event: threading.Event,
    metadatos_ref: list,
    done_event: threading.Event,
) -> None:
    """
    Coordina el procesamiento paralelo.
    - Filtra archivos ya procesados (reanudación automática).
    - Persiste progreso después de cada sentencia.
    - Este hilo NUNCA modifica st.session_state (thread-safe con Streamlit).
    - Señaliza done_event al terminar (completo o pausado).
    """
    fecha_inicio = datetime.now().isoformat()
    ts_inicio    = time.time()
    progreso     = cargar_progreso()
    n_ya_procesados = len(progreso)   # para calcular velocidad real de esta sesión
    escribir_estado("Iniciando análisis...", inicio=ts_inicio, n_previo=n_ya_procesados)

    nombres_ordenados = sorted(archivos_dict.keys())
    hash_corpus = hashlib.md5("".join(nombres_ordenados).encode("utf-8")).hexdigest()

    pendientes = {
        nombre: texto
        for nombre, texto in archivos_dict.items()
        if nombre not in progreso
    }

    def tarea(nombre: str, texto: str):
        return nombre, analizar_sentencia(nombre, texto, api_key, pausa_event)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futuros = {
            executor.submit(tarea, nombre, texto): nombre
            for nombre, texto in pendientes.items()
        }

        for futuro in as_completed(futuros):
            if pausa_event.is_set():
                for f in futuros:
                    f.cancel()
                break

            try:
                nombre, resultado = futuro.result()
                progreso[nombre] = resultado
                guardar_progreso(progreso)
            except Exception:
                nombre = futuros[futuro]
                progreso[nombre] = _resultado_error(nombre)
                guardar_progreso(progreso)

    # Limpiar estado al terminar
    escribir_estado("")

    # Exportaciones finales (solo si no se pausó)
    if not pausa_event.is_set():
        fecha_fin = datetime.now().isoformat()
        metadatos = {
            "fecha_inicio": fecha_inicio,
            "fecha_fin":    fecha_fin,
            "hash_corpus":  hash_corpus,
            "total_corpus": len(archivos_dict),
        }
        metadatos_ref.append(metadatos)

        try:
            bitacora = generar_bitacora(progreso, metadatos)
            with open(ARCHIVO_BITACORA, "w", encoding="utf-8") as f:
                json.dump(bitacora, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    done_event.set()


# =============================================================
# MÓDULO 5 — EXPORTACIÓN
# =============================================================

def generar_excel(progreso: dict) -> bytes:
    """
    Genera el Excel con dos hojas:
      - Resultados: una fila por sentencia, columnas de validación en amarillo.
      - Resumen:    frecuencias por patrón.
    Devuelve bytes para descarga directa desde Streamlit.
    """
    wb = openpyxl.Workbook()

    # ── Estilos ────────────────────────────────────────────────────────────
    fill_encabezado = PatternFill(fill_type="solid", fgColor="D9D9D9")
    fill_amarillo   = PatternFill(fill_type="solid", fgColor="FFFF00")
    fill_gris_claro = PatternFill(fill_type="solid", fgColor="F5F5F5")
    fill_rojo       = PatternFill(fill_type="solid", fgColor="FFCCCC")
    fuente_negrita  = Font(bold=True)
    fuente_nota     = Font(italic=True, color="666666")

    # ── Hoja 1: Resultados ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resultados"

    # Construir definición de columnas dinámicamente
    prefijos_patron = [
        ("p1", "p1_omision"),
        ("p2", "p2_precariedad"),
        ("p3", "p3_documento_digital"),
        ("p4", "p4_sobreseimiento"),
        ("p5", "p5_no_transcripcion"),
    ]

    cols_def = [("archivo", 35, False)]  # (nombre_col, ancho, es_validacion)
    for pref, _ in prefijos_patron:
        cols_def.append((f"{pref}_presente",   12, False))
        cols_def.append((f"{pref}_fragmento",  60, False))
        cols_def.append((f"{pref}_validacion", 15, True))
    cols_def.append(("observaciones", 30, False))

    # Índices base-1 de columnas de validación (para no pisar amarillo con fill de fila)
    idx_validacion = set()
    for idx, (_, _, es_val) in enumerate(cols_def, start=1):
        if es_val:
            idx_validacion.add(idx)

    # Escribir encabezados
    for idx, (nombre_col, ancho, es_val) in enumerate(cols_def, start=1):
        celda = ws.cell(row=1, column=idx, value=nombre_col)
        celda.font      = fuente_negrita
        celda.fill      = fill_amarillo if es_val else fill_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    # Escribir filas de datos
    for fila, (nombre_archivo, resultado) in enumerate(sorted(progreso.items()), start=2):
        es_error  = resultado.get("_error", False)
        algun_si  = any(resultado.get(p, {}).get("presente", False) for p in PATRONES_LISTA)

        ws.cell(row=fila, column=1, value=nombre_archivo).alignment = Alignment(vertical="top")

        col = 2
        for _, patron in prefijos_patron:
            datos_p   = resultado.get(patron, {"presente": False, "fragmento": ""})
            presente  = datos_p.get("presente", False)
            fragmento = datos_p.get("fragmento", "")

            c_pres = ws.cell(row=fila, column=col, value="Sí" if presente else "No")
            c_pres.alignment = Alignment(horizontal="center", vertical="top")
            col += 1

            c_frag = ws.cell(row=fila, column=col, value=fragmento)
            c_frag.alignment = Alignment(vertical="top", wrap_text=True)
            col += 1

            ws.cell(row=fila, column=col, value="").fill = fill_amarillo
            col += 1

        ws.cell(row=fila, column=col, value="")  # observaciones

        # Color y altura según tipo de fila
        if es_error:
            fill_fila = fill_rojo
            altura    = 20
        elif algun_si:
            fill_fila = None   # blanco
            altura    = 60
        else:
            fill_fila = fill_gris_claro
            altura    = 20

        ws.row_dimensions[fila].height = altura

        if fill_fila:
            for c in range(1, col + 1):
                if c not in idx_validacion:
                    ws.cell(row=fila, column=c).fill = fill_fila

    # ── Hoja 2: Resumen ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")

    enc_res    = ["Clave", "Nombre del patrón", "Total detectados", "% del corpus procesado"]
    anchos_res = [8, 55, 18, 25]

    for i, (enc, ancho) in enumerate(zip(enc_res, anchos_res), start=1):
        c = ws2.cell(row=1, column=i, value=enc)
        c.font = fuente_negrita
        c.fill = fill_encabezado
        ws2.column_dimensions[get_column_letter(i)].width = ancho

    nombres_largos = {
        "p1_omision":          "Omisión administrativa",
        "p2_precariedad":      "Cláusula de precariedad",
        "p3_documento_digital":"Rechazo de documento digital",
        "p4_sobreseimiento":   "Sesgo de anclaje en sobreseimiento",
        "p5_no_transcripcion": "No transcripción de conceptos de violación",
    }
    claves_cortas = {
        "p1_omision": "P1", "p2_precariedad": "P2", "p3_documento_digital": "P3",
        "p4_sobreseimiento": "P4", "p5_no_transcripcion": "P5",
    }

    total_proc = len([r for r in progreso.values() if r.get("_procesado")])

    for fila_res, patron in enumerate(PATRONES_LISTA, start=2):
        total_det = sum(1 for r in progreso.values() if r.get(patron, {}).get("presente", False))
        pct = (total_det / total_proc * 100) if total_proc > 0 else 0.0

        ws2.cell(row=fila_res, column=1, value=claves_cortas[patron])
        ws2.cell(row=fila_res, column=2, value=nombres_largos[patron])
        ws2.cell(row=fila_res, column=3, value=total_det)
        ws2.cell(row=fila_res, column=4, value=f"{pct:.1f}%")

    # Nota al pie
    nota_fila = len(PATRONES_LISTA) + 3
    c_nota = ws2.cell(
        row=nota_fila, column=1,
        value=(
            "Nota: Los porcentajes corresponden a detección automática de la IA, "
            "pendiente de validación manual del investigador."
        ),
    )
    c_nota.font = fuente_nota
    ws2.merge_cells(f"A{nota_fila}:D{nota_fila}")

    # Serializar a bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generar_bitacora(progreso: dict, metadatos: dict) -> dict:
    """Construye el JSON de bitácora metodológica completo."""
    total_proc    = len([r for r in progreso.values() if r.get("_procesado")])
    total_errores = len([r for r in progreso.values() if r.get("_error")])

    resultados_por_patron = {}
    for patron in PATRONES_LISTA:
        total = sum(1 for r in progreso.values() if r.get(patron, {}).get("presente", False))
        pct   = (total / total_proc * 100) if total_proc > 0 else 0.0
        resultados_por_patron[patron] = {"total": total, "porcentaje": round(pct, 2)}

    return {
        "fecha_inicio":            metadatos.get("fecha_inicio", ""),
        "fecha_fin":               metadatos.get("fecha_fin", ""),
        "modelo":                  MODELO,
        "total_sentencias_corpus": metadatos.get("total_corpus", 0),
        "sentencias_procesadas":   total_proc,
        "sentencias_con_error":    total_errores,
        "hash_corpus":             metadatos.get("hash_corpus", ""),
        "prompt_utilizado":        PROMPT_BASE,
        "patrones_definidos":      PATRONES_DESCRIPCION,
        "resultados_por_patron":   resultados_por_patron,
    }


# =============================================================
# MÓDULO 1 — INTERFAZ STREAMLIT
# =============================================================

def _init_estado() -> None:
    """Inicializa session_state con valores por defecto (solo la primera vez)."""
    defaults = {
        "corriendo":     False,
        "pausado":       False,
        "hilo":          None,
        "pausa_event":   threading.Event(),
        "done_event":    threading.Event(),
        "metadatos_ref": [],
        "archivos":      {},        # {nombre_archivo: contenido_str}
        "excel_bytes":   None,
        "bitacora_dict": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def main() -> None:
    st.set_page_config(
        page_title="Analizador de Amparos",
        page_icon="⚖️",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    _init_estado()

    st.title("⚖️ Analizador de Corpus de Amparos")
    st.caption(
        "Detección automática de patrones jurídicos · "
        "Investigación doctoral en derecho empírico"
    )
    st.divider()

    # ──────────────────────────────────────────────────────────
    # SECCIÓN 1 · CONFIGURACIÓN
    # ──────────────────────────────────────────────────────────
    st.subheader("1 · Configuración")
    col_key, col_files = st.columns([1, 2])

    with col_key:
        api_key = st.text_input(
            "Clave API de Anthropic",
            type="password",
            placeholder="sk-ant-...",
            disabled=st.session_state.corriendo,
            help="Su clave personal de Anthropic. No se almacena ni se comparte.",
        )

    with col_files:
        subidos = st.file_uploader(
            "Cargar sentencias (.txt)",
            type=["txt"],
            accept_multiple_files=True,
            disabled=st.session_state.corriendo,
            help="Puede seleccionar todos los archivos al mismo tiempo.",
        )

    # Almacenar archivos en session_state cuando se cargan
    if subidos:
        nuevos = {}
        for f in subidos:
            try:
                nuevos[f.name] = f.read().decode("utf-8", errors="replace")
            except Exception:
                pass
        if nuevos:
            st.session_state.archivos = nuevos

    n_archivos = len(st.session_state.archivos)
    if n_archivos > 0:
        st.success(f"📂 {n_archivos} sentencia(s) cargada(s) y listas para analizar.")

    # ──────────────────────────────────────────────────────────
    # AVISO DE REANUDACIÓN
    # ──────────────────────────────────────────────────────────
    progreso_disco   = cargar_progreso()
    n_ya_procesados  = len(progreso_disco)

    if (
        n_ya_procesados > 0
        and not st.session_state.corriendo
        and not st.session_state.pausado
        and st.session_state.excel_bytes is None
    ):
        st.divider()
        st.warning(
            f"⚠️ Se encontró un análisis anterior con **{n_ya_procesados}** "
            f"sentencia(s) procesada(s). ¿Desea continuar o empezar de nuevo?"
        )
        ca, cb = st.columns(2)
        with ca:
            if st.button(
                "▶️ Continuar desde donde quedó", use_container_width=True, type="primary"
            ):
                st.rerun()
        with cb:
            if st.button("🔄 Empezar de nuevo (borrar avance)", use_container_width=True):
                for archivo_borrar in [ARCHIVO_PROGRESO, ARCHIVO_BITACORA]:
                    if os.path.exists(archivo_borrar):
                        os.remove(archivo_borrar)
                st.session_state.excel_bytes  = None
                st.session_state.bitacora_dict = None
                st.rerun()

    # ──────────────────────────────────────────────────────────
    # SECCIÓN 2 · CONTROLES
    # ──────────────────────────────────────────────────────────
    st.divider()
    st.subheader("2 · Análisis")

    st.warning(
        "⚠️ **Importante:** No permita que la computadora entre en modo de **hibernación o suspensión** "
        "mientras el análisis está corriendo — interrumpe el proceso. \n\n"
        "Si la computadora se durmió o el programa se detuvo inesperadamente, "
        "**simplemente vuelva a cargar los archivos y presione ▶️ Reanudar** — "
        "el programa retomará desde la última sentencia guardada automáticamente.",
        icon="💤",
    )

    c1, c2, _ = st.columns([1, 1, 2])

    with c1:
        label_btn = "▶️ Reanudar análisis" if st.session_state.pausado else "▶️ Iniciar análisis"
        if st.button(
            label_btn,
            disabled=(not api_key or n_archivos == 0 or st.session_state.corriendo),
            use_container_width=True,
            type="primary",
        ):
            st.session_state.pausa_event  = threading.Event()
            st.session_state.done_event   = threading.Event()
            st.session_state.metadatos_ref = []
            st.session_state.corriendo    = True
            st.session_state.pausado      = False

            hilo = threading.Thread(
                target=correr_analisis,
                args=(
                    st.session_state.archivos,
                    api_key,
                    st.session_state.pausa_event,
                    st.session_state.metadatos_ref,
                    st.session_state.done_event,
                ),
                daemon=True,
            )
            st.session_state.hilo = hilo
            hilo.start()
            st.rerun()

    with c2:
        if st.button(
            "⏸️ Pausar",
            disabled=not st.session_state.corriendo,
            use_container_width=True,
        ):
            st.session_state.pausa_event.set()
            st.session_state.corriendo = False
            st.session_state.pausado   = True
            st.rerun()

    # ── Botón de reinicio ──────────────────────────────────────
    progreso_para_reset = cargar_progreso()
    n_proc_reset  = len(progreso_para_reset)
    n_err_reset   = sum(1 for r in progreso_para_reset.values() if r.get("_error"))

    if n_proc_reset > 0 and not st.session_state.corriendo:
        with st.expander("⚠️ Opciones avanzadas"):
            st.warning(
                f"Actualmente hay **{n_proc_reset} sentencias registradas** "
                f"({n_err_reset} con error). "
                "Si desea volver a analizarlas todas desde cero, use los botones de abajo."
            )
            col_r1, col_r2 = st.columns(2)

            with col_r1:
                if st.button(
                    "🔄 Reiniciar solo errores",
                    use_container_width=True,
                    help="Elimina solo las sentencias marcadas con error para que sean reintentadas",
                ):
                    progreso_limpio = {
                        k: v for k, v in progreso_para_reset.items()
                        if not v.get("_error", False)
                    }
                    guardar_progreso(progreso_limpio)
                    st.session_state.excel_bytes   = None
                    st.session_state.bitacora_dict = None
                    st.session_state.pausado       = False
                    st.success(f"✅ Se eliminaron {n_err_reset} errores. Presione ▶️ Reanudar para reprocesarlos.")
                    st.rerun()

            with col_r2:
                confirmacion = st.checkbox("Confirmo que quiero borrar TODO el progreso")
                if st.button(
                    "🗑️ Reiniciar desde cero",
                    use_container_width=True,
                    disabled=not confirmacion,
                    help="Borra todo el progreso y vuelve a analizar las 951 sentencias completas",
                ):
                    for archivo in [ARCHIVO_PROGRESO, ARCHIVO_ESTADO, ARCHIVO_BITACORA]:
                        try:
                            if os.path.exists(archivo):
                                os.remove(archivo)
                        except Exception:
                            pass
                    st.session_state.excel_bytes   = None
                    st.session_state.bitacora_dict = None
                    st.session_state.pausado       = False
                    st.success("✅ Progreso eliminado. Cargue los archivos y presione ▶️ Iniciar análisis.")
                    st.rerun()

    # ──────────────────────────────────────────────────────────
    # SECCIÓN 3 · PROGRESO EN TIEMPO REAL
    # ──────────────────────────────────────────────────────────
    progreso_actual = cargar_progreso()
    n_proc    = len(progreso_actual)
    n_total   = n_archivos if n_archivos > 0 else max(n_proc, 1)
    n_errores = sum(1 for r in progreso_actual.values() if r.get("_error"))

    if st.session_state.corriendo or st.session_state.pausado or n_proc > 0:
        st.divider()
        st.subheader("3 · Progreso")

        pct = min(n_proc / n_total, 1.0) if n_total > 0 else 0.0
        st.progress(pct, text=f"Procesadas: {n_proc} de {n_total}")

        m1, m2, m3 = st.columns(3)
        m1.metric("✅ Procesadas", n_proc)
        m2.metric("❌ Con error",  n_errores)
        m3.metric("⏳ Pendientes", max(n_total - n_proc, 0))

        if st.session_state.pausado:
            st.info("⏸️ El análisis está pausado. Presione **Reanudar** para continuar.")

        # Mostrar estado actual del hilo + ETA
        if st.session_state.corriendo:
            estado_dict = leer_estado()
            msg_estado  = estado_dict.get("mensaje", "")
            ts_inicio   = estado_dict.get("inicio")
            n_previo    = estado_dict.get("n_previo", 0)  # docs ya procesados al iniciar sesión

            # ETA — velocidad real = solo sentencias procesadas EN ESTA SESIÓN
            n_esta_sesion = max(n_proc - n_previo, 0)
            if ts_inicio and n_esta_sesion > 0:
                elapsed        = time.time() - ts_inicio
                tasa           = n_esta_sesion / elapsed   # sentencias/segundo reales
                pendientes_num = max(n_total - n_proc, 0)
                eta_sec        = pendientes_num / tasa if tasa > 0 else 0
                eta_min        = eta_sec / 60
                vel_min        = tasa * 60

                eta_col1, eta_col2 = st.columns(2)
                eta_col1.metric("⚡ Velocidad", f"{vel_min:.1f} sent/min")
                eta_col2.metric(
                    "🕐 Tiempo estimado restante",
                    f"~{eta_min:.0f} min" if eta_min > 1 else "< 1 min",
                )

            if msg_estado:
                if "Límite de API" in msg_estado or "Esperando" in msg_estado:
                    st.warning(f"🕐 {msg_estado}")
                elif "Error" in msg_estado or "Sin conexión" in msg_estado:
                    st.error(f"⚠️ {msg_estado}")
                elif msg_estado:
                    st.info(f"⚙️ {msg_estado}")

        # Monitorear si el hilo terminó
        if st.session_state.corriendo:
            if st.session_state.done_event.is_set():
                st.session_state.corriendo = False

                progreso_final = cargar_progreso()
                st.session_state.excel_bytes = generar_excel(progreso_final)

                if st.session_state.metadatos_ref:
                    bitacora = generar_bitacora(
                        progreso_final, st.session_state.metadatos_ref[0]
                    )
                    st.session_state.bitacora_dict = bitacora

                st.rerun()
            else:
                # Polling: esperar 1 segundo y volver a renderizar
                time.sleep(1.0)
                st.rerun()

    # ──────────────────────────────────────────────────────────
    # SECCIÓN 4 · RESULTADOS Y DESCARGA
    # ──────────────────────────────────────────────────────────
    progreso_final2 = cargar_progreso()
    hay_resultados  = len(progreso_final2) > 0 and not st.session_state.corriendo

    if hay_resultados:
        st.divider()
        st.subheader("4 · Resultados")

        nombres_largos_ui = {
            "p1_omision":          "Omisión administrativa",
            "p2_precariedad":      "Cláusula de precariedad",
            "p3_documento_digital":"Rechazo de documento digital",
            "p4_sobreseimiento":   "Sesgo de anclaje en sobreseimiento",
            "p5_no_transcripcion": "No transcripción de conceptos de violación",
        }
        claves_cortas_ui = {
            "p1_omision": "P1", "p2_precariedad": "P2", "p3_documento_digital": "P3",
            "p4_sobreseimiento": "P4", "p5_no_transcripcion": "P5",
        }

        total_proc_ui = len([r for r in progreso_final2.values() if r.get("_procesado")])

        filas_resumen = []
        for patron in PATRONES_LISTA:
            total_det = sum(
                1 for r in progreso_final2.values()
                if r.get(patron, {}).get("presente", False)
            )
            pct = (total_det / total_proc_ui * 100) if total_proc_ui > 0 else 0.0
            filas_resumen.append({
                "Clave":       claves_cortas_ui[patron],
                "Patrón":      nombres_largos_ui[patron],
                "Detectadas":  total_det,
                "% del corpus":f"{pct:.1f}%",
            })

        df_res = pd.DataFrame(filas_resumen)
        st.dataframe(df_res, use_container_width=True, hide_index=True)
        st.caption(
            "⚠️ Los resultados son detección automática de la IA. "
            "Cada hallazgo positivo debe ser validado manualmente por el investigador."
        )

        # Botones de descarga
        st.markdown("**Descargar resultados:**")
        dc1, dc2 = st.columns(2)

        with dc1:
            if st.session_state.excel_bytes is None and total_proc_ui > 0:
                st.session_state.excel_bytes = generar_excel(progreso_final2)

            if st.session_state.excel_bytes:
                st.download_button(
                    label="📥 Excel de resultados (.xlsx)",
                    data=st.session_state.excel_bytes,
                    file_name=ARCHIVO_EXCEL,
                    mime=(
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    ),
                    use_container_width=True,
                    type="primary",
                )

        with dc2:
            if st.session_state.bitacora_dict:
                st.download_button(
                    label="📋 Bitácora metodológica (.json)",
                    data=json.dumps(
                        st.session_state.bitacora_dict, ensure_ascii=False, indent=2
                    ),
                    file_name=ARCHIVO_BITACORA,
                    mime="application/json",
                    use_container_width=True,
                )

    # ──────────────────────────────────────────────────────────
    # SECCIÓN 5 · VALIDACIÓN DE HALLAZGOS
    # ──────────────────────────────────────────────────────────
    progreso_final3 = cargar_progreso()
    sentencias_con_positivos = {
        nombre: res
        for nombre, res in progreso_final3.items()
        if any(res.get(p, {}).get("presente", False) for p in PATRONES_LISTA)
    }

    if sentencias_con_positivos:
        st.divider()
        st.subheader("5 · Validación de hallazgos")
        st.caption(
            "Revise cada hallazgo positivo. Haga clic en el patrón para ver el fragmento "
            "y el documento original. Valide con los botones de radio."
        )

        nombres_largos_val = {
            "p1_omision":          "Omisión administrativa",
            "p2_precariedad":      "Cláusula de precariedad",
            "p3_documento_digital":"Rechazo de documento digital",
            "p4_sobreseimiento":   "Sesgo de anclaje en sobreseimiento",
            "p5_no_transcripcion": "No transcripción de conceptos de violación",
        }

        claves_cortas_val = {
            "p1_omision": "P1", "p2_precariedad": "P2", "p3_documento_digital": "P3",
            "p4_sobreseimiento": "P4", "p5_no_transcripcion": "P5",
        }

        # Mostrar lista de sentencias con positivos
        for nombre_archivo, resultado in sorted(sentencias_con_positivos.items()):
            patrones_positivos = [
                p for p in PATRONES_LISTA
                if resultado.get(p, {}).get("presente", False)
            ]

            if not patrones_positivos:
                continue

            # Crear expander para esta sentencia
            cantidad = len(patrones_positivos)
            label_expander = f"📄 {nombre_archivo} ({cantidad} patrón{'es' if cantidad > 1 else ''})"

            with st.expander(label_expander, expanded=False):
                # Para cada patrón positivo en esta sentencia
                for patron in patrones_positivos:
                    datos_patron = resultado.get(patron, {})
                    fragmento = datos_patron.get("fragmento", "")

                    # Encabezado con clave y nombre
                    col_nombre, col_info = st.columns([0.7, 0.3])
                    col_nombre.markdown(
                        f"**{claves_cortas_val[patron]} — {nombres_largos_val[patron]}**"
                    )

                    # Obtener estado de validación previo
                    clave_val = f"_validacion_{patron}"
                    estado_actual = resultado.get(clave_val, "Pendiente")

                    # Radio buttons para validación
                    with col_info:
                        estado_nuevo = st.radio(
                            label="Estado",
                            options=["Pendiente", "✅ Confirmado", "❌ Rechazado"],
                            index=0 if estado_actual == "Pendiente" else (1 if estado_actual == "✅ Confirmado" else 2),
                            key=f"validacion_{nombre_archivo}_{patron}",
                            horizontal=True,
                        )

                        # Guardar si cambió
                        if estado_nuevo != estado_actual:
                            guardar_validacion(nombre_archivo, patron, estado_nuevo)

                    # Mostrar fragmento en área de texto (deshabilitada)
                    st.text_area(
                        label=f"Fragmento identificado:",
                        value=fragmento,
                        height=100,
                        disabled=True,
                        key=f"fragmento_{nombre_archivo}_{patron}",
                    )

                    # Expandible para ver documento completo
                    if "archivos" in st.session_state and st.session_state.archivos:
                        with st.expander("📑 Ver documento original completo"):
                            # Buscar el archivo en archivos cargados
                            texto_original = None
                            for nom, txt in st.session_state.archivos.items():
                                if nom == nombre_archivo:
                                    texto_original = txt
                                    break

                            if texto_original:
                                st.text_area(
                                    label="Documento completo:",
                                    value=texto_original,
                                    height=300,
                                    disabled=True,
                                    key=f"doc_completo_{nombre_archivo}_{patron}",
                                )
                            else:
                                st.warning("⚠️ El archivo original no está cargado. Recargue los archivos en la Sección 1.")
                    else:
                        st.info("ℹ️ Cargue archivos en la Sección 1 para ver documentos originales aquí.")

                    st.divider()

    # ──────────────────────────────────────────────────────────
    # PIE DE PÁGINA
    # ──────────────────────────────────────────────────────────
    st.divider()
    st.caption(
        f"Modelo: {MODELO}  ·  "
        "Sistema para uso exclusivo de investigación doctoral  ·  "
        "Versión 1.0"
    )


if __name__ == "__main__":
    main()
